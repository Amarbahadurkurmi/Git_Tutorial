# *********************************************************************************************
from tkinter import *
from tkinter import messagebox,ttk,Menu
import tkinter as tk
import os
import openpyxl as xl
import pandas as pd
from datetime import datetime,timedelta
import menu

########################## Start Main Window Programm ############################################
global acpcs,acdcs,fcpcs,fcdcs,iipcs,mepcs,medcs,smcs,ltcs,staff,wd,tcs,mnth,yer
global acdamt,acpamt,fcpamt,fcdamt,iipamt,mepamt,medamt,tamt,cmb,subcslbl,cmbox

win = tk.Tk()
win.title("Staion PCDO")
width= win.winfo_screenwidth()               
height= win.winfo_screenheight()               
win.geometry("%dx%d" % (width, height))

menu.menub(root=win,tk=tk)


################################ Staion selcection command #####################################
def validate_cmbox(event):
    df = pd.read_excel('stnpcdo.xlsx')  # Make sure your file path is correct

    enter_stn_list = df.iloc[:,0].tolist()

    for i in enter_stn_list:
        if i==cmbox.get():
            messagebox.showerror('Error','Station alredy entered')
        

############### Mthod for month and year ###############################

def get_display_month():
    today = datetime.today()
    if today.day == 1:
        # Move to the last day of the previous month
        previous_month = today.replace(day=1) - timedelta(days=1)
        display_month = previous_month.strftime("%B")
    else:
        display_month = today.strftime("%B")
    return display_month

def get_display_year():
    today = datetime.today()
    if today.day == 1:
        # Move to the last day of the previous month
        previous_year = today.replace(month=1) - timedelta(month=1)
        display_year = previous_month.strftime("%Y")
    else:
        display_year = today.strftime("%Y")
    return display_year
###################################################################################################

def insert_data():
#     # Get data from Entry widgets 
    if (stncomb.get()==''):
        messagebox.showerror('Error','Please Select station')
    elif((totalcs.cget('text')=='' or 0) and (totalamt.cget('text')=='' or 0)):
        messagebox.showerror('Error','Please Enter Minimum 1 C/S and Amount ')
    elif((staff.get()=='' or 0) and (wd.get()==' or 0')):
        messagebox.showerror('Error','Please Enter staff and working days')
    else:   
        list1 = cmbox.get()
        acs= int(acpcs.get())
        fcs = int(fcpcs.get())
        iics = int(iipcs.get())
        ucs = int(ublcs.get())
        mec = int(mepcs.get())
        acdc = int(acdcs.get())
        fcdc = int(fcdcs.get())
        medc = int(medcs.get())
        stf = int(staff.get())
        w = int(wd.get())
        ltc = int(ltcs.get())
        smc = int(smcs.get())
        
        tcs= totalcs.cget('text')
        
        aca= int(acpamt.get())
        fca = int(fcpamt.get())
        iia = int(iipamt.get())
        ua = int(ublamt.get())
        mea = int(mepamt.get())
        acda = int(acdamt.get())
        fcda = int(fcdamt.get())
        meda = int(medamt.get())
        lta = int(ltamt.get())
        sma =int(smamt.get())
        prd = pradio.get()
        tamt = totalamt.cget('text')
        mnth = month.cget('text')
        yer = year.cget('text')
        
    #     # # Load the Excel file or create a new one if it doesn't exist
        
        try:
            workbook = xl.load_workbook('stnpcdo.xlsx')
        except FileNotFoundError:
            workbook = xl.Workbook()
            workbook.remove(workbook.active)  # Remove the default sheet
            sheet = workbook.create_sheet(title='Sheet1')
            sheet.append(['STN', 'AC_PWT_CS', 'AC_PWT_AMT','AC_DIFF_CS', 'AC_DIFF_AMT',
                        'FC_PWT_CS','FC_DIFF_AMT', 'FC_DIFF_CS','FC_PWT_AMT',
                        'II_PWT_CS','II_PWT_AMT', 'UBL_CS','UBL_AMT','TOTAL_CS','TOTAL_AMT',
                        'STAFF','WD','LITT_CS','LITT_AMT', 'SM_CS','SM_AMT',
                        'ME_PWT_CS', 'ME_PWT_AMT','ME_DIFF_CS', 'ME_DIFF_AMT','PERIOD','MONTH','YEAR'])  # Add headers
        else:
            sheet = workbook.active

        
            # Append the new data
        sheet.append([list1,acs,aca,acdc,acda,fcs,fca,fcdc,fcda,iics,iia,ucs,ua,tcs,tamt,stf,w,ltc,lta,
                    smc,sma,mec,mea,medc,meda,prd,mnth,yer])
        
        # # Save the Excel file
        workbook.save('stnpcdo.xlsx')
        clear()
        update_summery()
        
        messagebox.showinfo("Success", "Data inserted successfully!")


################################# Clear command function ############################################
def clear():
    acpcs.delete(0,END)
    fcpcs.delete(0,END)
    iipcs.delete(0,END)
    ublcs.delete(0,END)
    mepcs.delete(0,END)
    acdcs.delete(0,END)
    fcdcs.delete(0,END)
    medcs.delete(0,END)
    ltcs.delete(0,END)
    smcs.delete(0,END)
    staff.delete(0,END)
    wd.delete(0,END)
    totalcs.config(text='')
    
    # Amount
    acpamt.delete(0,END)
    fcpamt.delete(0,END)
    iipamt.delete(0,END)
    ublamt.delete(0,END)
    mepamt.delete(0,END)
    acdamt.delete(0,END)
    fcdamt.delete(0,END)
    medamt.delete(0,END)
    ltamt.delete(0,END)
    smamt.delete(0,END)
    totalamt.config(text='')
    set_zero()
    stncomb.focus()

################################# Reset and ste Zeo in entry wigdet ############################################

def set_zero():

    acpcs.insert(0,0)
    acpamt.insert(0,0)
    fcpcs.insert(0,0)
    fcpamt.insert(0,0)
    iipcs.insert(0,0)
    iipamt.insert(0,0)
    ublcs.insert(0,0)
    ublamt.insert(0,0)
    mepcs.insert(0,0)
    mepamt.insert(0,0)
    medcs.insert(0,0)
    medamt.insert(0,0)
    staff.insert(0,0)
    wd.insert(0,0)
    ltcs.insert(0,0)
    ltamt.insert(0,0)
    smcs.insert(0,0)
    smamt.insert(0,0)
    acdcs.insert(0,0)
    acdamt.insert(0,0)
    fcdcs.insert(0,0)
    fcdamt.insert(0,0)

########################### Caluculate the total C/s and Amt #########################################

def show(*args):
    # Cases
    cmb = cmbox.get()
    acs= int(acpcs.get()or 0)
    fcs = int(fcpcs.get()or 0)
    iics = int(iipcs.get() or 0)
    ucs = int(ublcs.get()or 0)
    mec = int(mepcs.get()or 0)
    acdc = int(acdcs.get()or 0)
    fcdc = int(fcdcs.get()or 0)
    medc = int(medcs.get()or 0)
    stf = int(staff.get()or 0)
    w = int(wd.get()or 0)
    ltc = int(ltcs.get()or 0)
    smc = int(smcs.get()or 0)
    
    tcs = acs+fcs+iics+mec+acdc+fcdc+medc+ucs
    
    totalcs.config(text=tcs)
    
    # Amount
    aca= int(acpamt.get()or 0)
    fca = int(fcpamt.get()or 0)
    iia = int(iipamt.get()or 0)
    ua = int(ublamt.get()or 0)
    mea = int(mepamt.get()or 0)
    acda = int(acdamt.get()or 0)
    fcda = int(fcdamt.get()or 0)
    meda = int(medamt.get()or 0)
    lta = int(ltamt.get()or 0)
    sma =int(smamt.get()or 0)
    
    
    tamt = aca+fca+iia+mea+acda+fcda+meda+ua
    
    totalamt.config(text=tamt)
    

def setradio():
    select_value = pradio.get()

################################# Tital label pack ############################################   

tk.Label(win,text='STATION PCDO ENTRY FORM',font=('New Times Roman',20,'bold'),relief=GROOVE,padx=10,
                      pady=1,bd=5,fg='dark slate blue',bg='sky blue').pack(fill=X,ipadx=5,ipady=2)


############################# Station selection and control pack ####################################

stnlable = ttk.Labelframe(win,)

tk.Label(stnlable,text='Station Name ',font=('Times New Roman',12,'bold'),
                     borderwidth=10,padx=20).grid(row=0,column=0)
cmbox= tk.StringVar()

df= pd.read_excel('stnlist.xlsx')
list1 = df.iloc[:,0].tolist()
stncomb = ttk.Combobox(stnlable,values=list1,textvariable=cmbox)
stncomb.grid(row=0,column=1,padx=10)
# stncomb.set(list1[0])
stncomb.bind("<<ComboboxSelected>>",validate_cmbox)

tk.Label(stnlable,text='Select Period',font=('Times New Roman',12,'bold')).grid(row=0,column=2,padx=10)

pradio = tk.IntVar()
radio=tk.Radiobutton(stnlable,text='I st ',font=('Times New Roman',12,'bold'),value=1,
                           variable=pradio,command=setradio).grid(row=0,column=3,padx=10)

radio=tk.Radiobutton(stnlable,text='II nd ',font=('Times New Roman',12,'bold'),value=2,
                           variable=pradio,).grid(row=0,column=4,padx=10)

radio=tk.Radiobutton(stnlable,text='III rd ',font=('Times New Roman',12,'bold'),value=3,
                           variable=pradio).grid(row=0,column=5,padx=10)
pradio.set(1)

addbtn = tk.Button(stnlable,text='ADD',command=insert_data,padx=10,pady=5,font=('Times New Roman',12,'bold'),
                      fg='brown',bg='light green').grid(row=0,column=6,padx=10,pady=10)


clrhbutton = tk.Button(stnlable,text='Clear',padx=10,pady=5,font=('Times New Roman',12,'bold'),
                      fg='black',bg='white',command=clear).grid(row=0,column=7,padx=10,pady=10)

month = tk.Label(stnlable,text=get_display_month(),font=('Times New Roman',15,'bold'),
                     padx=20,border=3,relief=FLAT,fg="brown")
month.grid(row=0,column=8)

year = tk.Label(stnlable,text=get_display_year(),font=('Times New Roman',15,'bold'),
                     padx=20,border=3,relief=FLAT,fg="brown")
year.grid(row=0,column=9)



stnlable.pack()
###################################### Main Entry pack #############################################
# Entry form
mainframe = tk.LabelFrame(win,text='Main Entry form',font=('New Times Roman',10,'bold'),
            fg='blue',padx=10,relief=GROOVE,border=5)

##############################################################################################################
#AC Local frame and Entry

acframe = tk.LabelFrame(mainframe,text='AC LOCAL',fg='green',font=('bold',10),relief=GROOVE,border=5)

tk.Label(acframe,text='PWT',font=('New Times Roman',12,'bold')).grid(row=0,column=0,columnspan=2)

tk.Label(acframe,text='C/S',font=('New Times Roman',12,'bold'),
                   relief=FLAT,padx=10,pady=10,bd=10).grid(row=1,column=0)

acpcs = tk.Entry(acframe,relief=GROOVE,width=10)
acpcs.grid(row=1,column=1)
acpcs.insert(0,0)
acpcs.bind("<KeyRelease>",show)

tk.Label(acframe,text='AMOUNT',font=('New Times Roman',12,'bold'),
                   relief=FLAT,padx=10,bd=10).grid(row=2,column=0)

acpamt = tk.Entry(acframe,relief=GROOVE,width=10)
acpamt.grid(row=2,column=1)
acpamt.insert(0,0)
acpamt.bind("<KeyRelease>",show)

tk.Label(acframe,text='Difference',font=('New Times Roman',12,'bold'),pady=5).grid(row=3,column=0,columnspan=2)

tk.Label(acframe,text='C/S',font=('New Times Roman',12,'bold'),
                   relief=FLAT,padx=10,pady=10,bd=10).grid(row=4,column=0)

acdcs = tk.Entry(acframe,relief=GROOVE,width=10)
acdcs.grid(row=4,column=1)
acdcs.insert(0,0)
acdcs.bind("<KeyRelease>",show)

tk.Label(acframe,text='AMOUNT',font=('New Times Roman',12,'bold'),
                   relief=FLAT,padx=10,bd=10).grid(row=5,column=0)

acdamt= tk.Entry(acframe,relief=GROOVE,width=10)
acdamt.grid(row=5,column=1)
acdamt.insert(0,0)
acdamt.bind("<KeyRelease>",show)

acframe.grid(row=0,column=0,padx=10,pady=5,ipadx=5) # AC Frame close

################################################################################################################
#FC Local frame and Entry

fcframe = tk.LabelFrame(mainframe,text='FC ',fg='red',font=('bold',10),relief=GROOVE,border=5) 

tk.Label(fcframe,text='PWT',font=('New Times Roman',12,'bold'),
                 ).grid(row=0,column=0,columnspan=2)

tk.Label(fcframe,text='C/S',font=('New Times Roman',12,'bold'),
                   relief=FLAT,padx=10,pady=10,bd=10).grid(row=1,column=0)

fcpcs = tk.Entry(fcframe,relief=GROOVE,width=10)
fcpcs.grid(row=1,column=1)
fcpcs.insert(0,0)
fcpcs.bind("<KeyRelease>",show)

tk.Label(fcframe,text='AMOUNT',font=('New Times Roman',12,'bold'),
                   relief=FLAT,padx=10,bd=10).grid(row=2,column=0)

fcpamt = tk.Entry(fcframe,relief=GROOVE,width=10)
fcpamt.grid(row=2,column=1)
fcpamt.insert(0,0)
fcpamt.bind("<KeyRelease>",show)

tk.Label(fcframe,text='Difference',font=('New Times Roman',12,'bold'),pady=5
                   ).grid(row=3,column=0,columnspan=2)

tk.Label(fcframe,text='C/S',font=('New Times Roman',12,'bold'),
                   relief=FLAT,padx=10,pady=10,bd=10).grid(row=4,column=0)

fcdcs = tk.Entry(fcframe,relief=GROOVE,width=10)
fcdcs.grid(row=4,column=1)
fcdcs.insert(0,0)
fcdcs.bind("<KeyRelease>",show)

tk.Label(fcframe,text='AMOUNT',font=('New Times Roman',12,'bold'),
                   relief=FLAT,padx=10,bd=10).grid(row=5,column=0)

fcdamt = tk.Entry(fcframe,relief=GROOVE,width=10)
fcdamt.grid(row=5,column=1)
fcdamt.insert(0,0)
fcdamt.bind("<KeyRelease>",show)

fcframe.grid(row=0,column=1,padx=10,pady=5,ipadx=5) # FC Frame close

#####################################################################################################################
#IInd Local frame and Entry

iiframe = tk.LabelFrame(mainframe,text='IInd Ord',fg='red',font=('bold',10),relief=GROOVE,border=5) 

tk.Label(iiframe,text='PWT',font=('New Times Roman',12,'bold')).grid(row=0,column=0,columnspan=2)

tk.Label(iiframe,text='C/S',font=('New Times Roman',12,'bold'),
                   relief=FLAT,padx=10,pady=10,bd=10).grid(row=1,column=0)

iipcs = tk.Entry(iiframe,relief=GROOVE,width=10)
iipcs.grid(row=1,column=1)
iipcs.insert(0,0)
iipcs.bind("<KeyRelease>",show)

tk.Label(iiframe,text='AMOUNT',font=('New Times Roman',12,'bold'),
                   relief=FLAT,padx=10,bd=10).grid(row=2,column=0)

iipamt = tk.Entry(iiframe,relief=GROOVE,width=10)
iipamt.grid(row=2,column=1)
iipamt.insert(0,0)
iipamt.bind("<KeyRelease>",show)

tk.Label(iiframe,text='UBL',font=('New Times Roman',12,'bold'),pady=5
         ).grid(row=3,column=0,columnspan=2)

tk.Label(iiframe,text='C/S',font=('New Times Roman',12,'bold'),
                   relief=FLAT,padx=10,pady=10,bd=10).grid(row=4,column=0)

ublcs = tk.Entry(iiframe,relief=GROOVE,width=10)
ublcs.grid(row=4,column=1)
ublcs.insert(0,0)
ublcs.bind("<KeyRelease>",show)

tk.Label(iiframe,text='AMOUNT',font=('New Times Roman',12,'bold'),
                   relief=FLAT,padx=10,bd=10).grid(row=5,column=0)

ublamt = tk.Entry(iiframe,relief=GROOVE,width=10)
ublamt.grid(row=5,column=1)
ublamt.insert(0,0)
ublamt.bind("<KeyRelease>",show)

iiframe.grid(row=0,column=2,padx=10,pady=5,ipadx=5) # iind orFrame close
###############################################################################################################
#Total Staff postion Local frame and Entry

totalframe = tk.LabelFrame(mainframe,text='TOTAL & STAFF POSITION',fg='green',font=('bold',10),relief=GROOVE,border=5) 

tk.Label(totalframe,text='Total',font=('New Times Roman',12,'bold')).grid(row=0,column=0,columnspan=2)

tk.Label(totalframe,text='C/S',font=('New Times Roman',12,'bold'),
                   relief=FLAT,padx=10,pady=10,bd=10).grid(row=1,column=0)

totalcs = tk.Label(totalframe,relief=GROOVE,width=10,font=('New Times Roman',12,'bold'),
                   fg='red')
totalcs.grid(row=1,column=1)
acpcs.bind("<KeyRelease>",show)

tk.Label(totalframe,text='AMOUNT',font=('New Times Roman',12,'bold'),
                   relief=FLAT,padx=10,bd=10).grid(row=2,column=0)

totalamt = tk.Label(totalframe,relief=GROOVE,width=10,font=('New Times Roman',12,'bold'),
                    fg='red')
totalamt.grid(row=2,column=1)

tk.Label(totalframe,text='Staff',font=('New Times Roman',12,'bold'),pady=5
                          ).grid(row=3,column=0,columnspan=2)

tk.Label(totalframe,text='STAFF',font=('New Times Roman',12,'bold'),
                   relief=FLAT,padx=10,pady=10,bd=10).grid(row=4,column=0)

staff = tk.Entry(totalframe,relief=GROOVE,width=10)
staff.grid(row=4,column=1)
staff.insert(0,0)

tk.Label(totalframe,text='W/D ',font=('New Times Roman',12,'bold'),
                   relief=FLAT,padx=10,bd=10).grid(row=5,column=0)

wd = tk.Entry(totalframe,relief=GROOVE,width=10)
wd.grid(row=5,column=1)
wd.insert(0,0)

totalframe.grid(row=0,column=3,padx=10,pady=10,ipadx=5) # UBL and Staff Frame close
##############################################################################################################
#Littering and Smoking Local frame and Entry

ltframe = tk.LabelFrame(mainframe,text='LITTERING & SMOKING',fg='red',font=('bold',10),relief=GROOVE,border=5,padx=10) 

tk.Label(ltframe,text='Littering',font=('New Times Roman',12,'bold')).grid(row=0,column=0,columnspan=2)

tk.Label(ltframe,text='C/S',font=('New Times Roman',12,'bold'),
                   relief=FLAT,padx=10,pady=10,bd=10).grid(row=1,column=0)

ltcs = tk.Entry(ltframe,relief=GROOVE,width=10)
ltcs.grid(row=1,column=1)
ltcs.insert(0,0)

tk.Label(ltframe,text='AMOUNT',font=('New Times Roman',12,'bold'),
                   relief=FLAT,padx=10,bd=10).grid(row=2,column=0)

ltamt = tk.Entry(ltframe,relief=GROOVE,width=10)
ltamt.grid(row=2,column=1)
ltamt.insert(0,0)

tk.Label(ltframe,text='Smoking',font=('New Times Roman',12,'bold'),pady=5).grid(row=3,column=0,columnspan=2)

tk.Label(ltframe,text='C/S',font=('New Times Roman',12,'bold'),
                   relief=FLAT,padx=10,pady=10,bd=10).grid(row=4,column=0)

smcs = tk.Entry(ltframe,relief=GROOVE,width=10)
smcs.grid(row=4,column=1)
smcs.insert(0,0)

tk.Label(ltframe,text='AMOUNT',font=('New Times Roman',12,'bold'),
                   relief=FLAT,padx=10,bd=10).grid(row=5,column=0)

smamt = tk.Entry(ltframe,relief=GROOVE,width=10)
smamt.grid(row=5,column=1)
smamt.insert(0,0)

ltframe.grid(row=0,column=4,padx=10) # LIttering and Smoking Frame close
#######################################################################################################
#Mail Express frame and Entry

meframe = tk.LabelFrame(mainframe,text='M/E',fg='green',font=('bold',10),relief=GROOVE,border=5) 

tk.Label(meframe,text='PWT',font=('New Times Roman',12,'bold')).grid(row=0,column=0,columnspan=2)

tk.Label(meframe,text='C/S',font=('New Times Roman',12,'bold'),
                   relief=FLAT,padx=10,pady=10,bd=10).grid(row=1,column=0)

mepcs = tk.Entry(meframe,relief=GROOVE,width=10)
mepcs.grid(row=1,column=1)
mepcs.insert(0,0)
mepcs.bind("<KeyRelease>",show)

tk.Label(meframe,text='AMOUNT',font=('New Times Roman',12,'bold'),
                   relief=FLAT,padx=10,bd=10).grid(row=2,column=0)

mepamt = tk.Entry(meframe,relief=GROOVE,width=10)
mepamt.grid(row=2,column=1)
mepamt.insert(0,0)
mepamt.bind("<KeyRelease>",show)

tk.Label(meframe,text='Difference',font=('New Times Roman',12,'bold'),
                   pady=5).grid(row=3,column=0,columnspan=2)

tk.Label(meframe,text='C/S',font=('New Times Roman',12,'bold'),
                   relief=FLAT,padx=10,pady=10,bd=10).grid(row=4,column=0)

medcs = tk.Entry(meframe,relief=GROOVE,width=10)
medcs.grid(row=4,column=1)
medcs.insert(0,0)
medcs.bind("<KeyRelease>",show)

tk.Label(meframe,text='AMOUNT',font=('New Times Roman',12,'bold'),
                   relief=FLAT,padx=10,bd=10).grid(row=5,column=0)

medamt = tk.Entry(meframe,relief=GROOVE,width=10)
medamt.grid(row=5,column=1)
medamt.insert(0,0)
medamt.bind("<KeyRelease>",show)

meframe.grid(row=0,column=5,padx=10,pady=10,ipadx=5) # ME Frame close

mainframe.pack()  #Main Frame close

# ################################# Summery frame pack ##################################################

summryframe = tk.LabelFrame(win,text='SUMMERY',font=('New Times Roman',10,'bold'),
            fg='blue',padx=10,relief=GROOVE,border=5)

suburbanframe = tk.LabelFrame(summryframe,text='Suburban',font=('New Times Roman',10,'bold'),
            fg='blue',padx=10,relief=GROOVE,border=5)

# Suburban Total

tk.Label(suburbanframe,text='Total C/S',font=('New Times Roman',12,'bold')).grid(row=0,column=0)

subcslbl=IntVar()
subcslbl= tk.Label(suburbanframe,text='',font=('New Times Roman',12,'bold'),
                   relief=GROOVE,padx=10,pady=5,bd=3,width=10)
subcslbl.grid(row=0,column=1)

tk.Label(suburbanframe,text='Total Amt',font=('New Times Roman',12,'bold')).grid(row=1,column=0)

subamtlbl=tk.Label(suburbanframe,text='',font=('New Times Roman',12,'bold'),
                   relief=GROOVE,padx=10,pady=5,bd=3,width=10)
subamtlbl.grid(row=1,column=1)

suburbanframe.grid(row=0,column=0,padx=2,ipady=5) # Suburban Close

# Main Line  Total

mainlineframe = tk.LabelFrame(summryframe,text='Mainline',font=('New Times Roman',10,'bold'),
            fg='blue',padx=10,relief=GROOVE,border=5)

tk.Label(mainlineframe,text='Total C/S',font=('New Times Roman',12,'bold')).grid(row=0,column=0)

mlcslbl=tk.Label(mainlineframe,text='',font=('New Times Roman',12,'bold'),
                   relief=GROOVE,padx=10,pady=5,bd=3,width=10)
mlcslbl.grid(row=0,column=1)

tk.Label(mainlineframe,text='Total Amt',font=('New Times Roman',12,'bold')).grid(row=1,column=0)

mlamtlbl=tk.Label(mainlineframe,text='',font=('New Times Roman',12,'bold'),
                   relief=GROOVE,padx=10,pady=5,bd=3,width=10)
mlamtlbl.grid(row=1,column=1)

mainlineframe.grid(row=0,column=1,padx=2,ipady=5) # Main line close

# UBL Total

ublframe = tk.LabelFrame(summryframe,text='UBL',font=('New Times Roman',10,'bold'),
            fg='blue',padx=10,relief=GROOVE,border=5)

tk.Label(ublframe,text='Total C/S',font=('New Times Roman',12,'bold')).grid(row=0,column=0)

ublcslbl=tk.Label(ublframe,text='',font=('New Times Roman',12,'bold'),
                   relief=GROOVE,padx=10,pady=5,bd=3,width=6)
ublcslbl.grid(row=0,column=1)

tk.Label(ublframe,text='Total Amt',font=('New Times Roman',12,'bold')).grid(row=1,column=0)

ublamtlbl=tk.Label(ublframe,text='',font=('New Times Roman',12,'bold'),
                   relief=GROOVE,padx=10,pady=5,bd=3,width=6)
ublamtlbl.grid(row=1,column=1)

ublframe.grid(row=0,column=2,padx=2,ipady=5) #UBL close

# Littering Total

litteringframe = tk.LabelFrame(summryframe,text='Littering',font=('New Times Roman',10,'bold'),
            fg='blue',padx=10,relief=GROOVE,border=5)

tk.Label(litteringframe,text='Total C/S',font=('New Times Roman',12,'bold')).grid(row=0,column=0)

ltcslbl=tk.Label(litteringframe,text='',font=('New Times Roman',12,'bold'),
                   relief=GROOVE,padx=10,pady=5,bd=3,width=6)
ltcslbl.grid(row=0,column=1)

tk.Label(litteringframe,text='Total Amt',font=('New Times Roman',12,'bold')).grid(row=1,column=0)

ltamtlbl=tk.Label(litteringframe,text='',font=('New Times Roman',12,'bold'),
                   relief=GROOVE,padx=10,pady=5,bd=3,width=6)
ltamtlbl.grid(row=1,column=1)

litteringframe.grid(row=0,column=3,padx=2,ipady=5) #Littering close

# smokint Total

smokingframe = tk.LabelFrame(summryframe,text='Smoking',font=('New Times Roman',10,'bold'),
            fg='blue',padx=10,relief=GROOVE,border=5)

tk.Label(smokingframe,text='Total C/S',font=('New Times Roman',12,'bold')).grid(row=0,column=0)

smcslbl=tk.Label(smokingframe,text='',font=('New Times Roman',12,'bold'),
                   relief=GROOVE,padx=10,pady=5,bd=3,width=5)
smcslbl.grid(row=0,column=1)

tk.Label(smokingframe,text='Total Amt',font=('New Times Roman',12,'bold')).grid(row=1,column=0)

smamtlbl=tk.Label(smokingframe,text='',font=('New Times Roman',12,'bold'),
                   relief=GROOVE,padx=10,pady=5,bd=3,width=5)
smamtlbl.grid(row=1,column=1)

smokingframe.grid(row=0,column=4,padx=2,ipady=5)  # Smoking close

# Grand  Total

grandtotal = tk.LabelFrame(summryframe,text='Grand Total',font=('New Times Roman',10,'bold'),
            fg='blue',padx=10,relief=GROOVE,border=5)

tk.Label(grandtotal,text='Total C/S',font=('New Times Roman',12,'bold')).grid(row=0,column=0)

gtcslbl=tk.Label(grandtotal,text='',font=('New Times Roman',12,'bold'),
                   relief=GROOVE,padx=10,pady=5,bd=3,width=12)
gtcslbl.grid(row=0,column=1)

tk.Label(grandtotal,text='Total Amt',font=('New Times Roman',12,'bold')).grid(row=1,column=0)

gtamtlbl=tk.Label(grandtotal,text='',font=('New Times Roman',12,'bold'),
                   relief=GROOVE,padx=10,pady=5,bd=3,width=12)
gtamtlbl.grid(row=1,column=1)

grandtotal.grid(row=0,column=5,padx=5,ipady=5) # Grand Total close

# slistframe = tk.LabelFrame(summryframe,text='STN Periodical Not received',font=('New Times Roman',10,'bold'),
#             fg='blue',padx=10,relief=GROOVE,border=5)

# slist = tk.Listbox(slistframe,width=20,height=5)
# slist.pack()

# countlist = tk.Label(slistframe,text='',width=10,bd=2,font=('New Times Roman',12,'bold'),
#                    relief=GROOVE,padx=10,pady=5)
# countlist.pack()


# slistframe.grid(row=0,column=5,padx=2)

summryframe.pack(ipadx=2,ipady=3) # Summery close

################################ AC / FC Bifercation ############################################

acfcframe = tk.Frame(win,border=5)

aclabelf=tk.LabelFrame(acfcframe,text='AC PWT',font=('New Times Roman',8,'bold'),fg='green')
tk.Label(aclabelf,text=' C/S',font=('New Times Roman',8,'bold')).grid(row=0,column=0)
acpcslabel=tk.Label(aclabelf,text='',font=('New Times Roman',8,'bold'),
                   relief=GROOVE,padx=10,pady=5,bd=3,width=10)
acpcslabel.grid(row=0,column=1)

tk.Label(aclabelf,text='Amount',font=('New Times Roman',8,'bold')).grid(row=0,column=2)
acpamtlabel=tk.Label(aclabelf,text='',font=('New Times Roman',8,'bold'),
                   relief=GROOVE,padx=10,pady=5,bd=3,width=10)
acpamtlabel.grid(row=0,column=3)
aclabelf.grid(row=0,column=0)
#_______________________________________________________________________________________________

acdlabelf=tk.LabelFrame(acfcframe,text='AC Diff',font=('New Times Roman',8,'bold'),fg='green')
tk.Label(acdlabelf,text=' C/S',font=('New Times Roman',8,'bold')).grid(row=0,column=0)
acdcslabel=tk.Label(acdlabelf,text='',font=('New Times Roman',8,'bold'),
                   relief=GROOVE,padx=10,pady=5,bd=3,width=10)
acdcslabel.grid(row=0,column=1)

tk.Label(acdlabelf,text='Amount',font=('New Times Roman',8,'bold')).grid(row=0,column=2)
acdamtlabel=tk.Label(acdlabelf,text='',font=('New Times Roman',8,'bold'),
                   relief=GROOVE,padx=10,pady=5,bd=3,width=10)
acdamtlabel.grid(row=0,column=3)
acdlabelf.grid(row=0,column=1)

############################### FC ##########################################

fclabelf=tk.LabelFrame(acfcframe,text='FC PWT',font=('New Times Roman',8,'bold'),fg='green')
tk.Label(fclabelf,text=' C/S',font=('New Times Roman',8,'bold')).grid(row=0,column=0)
fcpcslabel=tk.Label(fclabelf,text='',font=('New Times Roman',8,'bold'),
                   relief=GROOVE,padx=10,pady=5,bd=3,width=10)
fcpcslabel.grid(row=0,column=1)

tk.Label(fclabelf,text='Amount',font=('New Times Roman',8,'bold')).grid(row=0,column=2)
fcpamtlabel=tk.Label(fclabelf,text='',font=('New Times Roman',8,'bold'),
                   relief=GROOVE,padx=10,pady=5,bd=3,width=10)
fcpamtlabel.grid(row=0,column=3)
fclabelf.grid(row=0,column=2,padx=30)
#_______________________________________________________________________________________________

fcdlabelf=tk.LabelFrame(acfcframe,text='FC Diff',font=('New Times Roman',8,'bold'),fg='green')
tk.Label(fcdlabelf,text=' C/S',font=('New Times Roman',8,'bold')).grid(row=0,column=0)
fcdpcslabel=tk.Label(fcdlabelf,text='',font=('New Times Roman',8,'bold'),
                   relief=GROOVE,padx=10,pady=5,bd=3,width=10)
fcdpcslabel.grid(row=0,column=1)

tk.Label(fcdlabelf,text='Amount',font=('New Times Roman',8,'bold')).grid(row=0,column=2)
fcdamtlabel=tk.Label(fcdlabelf,text='',font=('New Times Roman',8,'bold'),
                   relief=GROOVE,padx=10,pady=5,bd=3,width=10)
fcdamtlabel.grid(row=0,column=3)
fcdlabelf.grid(row=0,column=3)


acfcframe.pack(ipadx=2,ipady=3) # Summery close


############################## fetch summery data ############################################

def update_summery():
    df = pd.read_excel('stnpcdo.xlsx')

    #  Suburban total 
    col1 = df.iloc[:,1]
    acpcslabel.config(text=col1.sum())
    col2 = df.iloc[:,3]
    acdcslabel.config(text=col2.sum())
    col3 = df.iloc[:,5]
    fcpcslabel.config(text=col3.sum())
    col4 = df.iloc[:,7]
    fcdpcslabel.config(text=col4.sum())
    col9 = df.iloc[:,9]
    subcslbl.config(text=col1.sum()+col2.sum()+col3.sum()+col4.sum()+col9.sum())

    col5 = df.iloc[:,2]
    acpamtlabel.config(text=col5.sum())
    col6 = df.iloc[:,4]
    acdamtlabel.config(text=col6.sum())
    col7 = df.iloc[:,6]
    fcpamtlabel.config(text=col7.sum())
    col8 = df.iloc[:,8]
    fcdamtlabel.config(text=col8.sum())
    col10 = df.iloc[:,10]
    subamtlbl.config(text=col5.sum()+col6.sum()+col7.sum()+col8.sum()+col10.sum())

    #  Main Line total 
    col21 = df.iloc[:,21]
    col23 = df.iloc[:,23]
    mlcslbl.config(text=col21.sum()+col23.sum())

    col22 = df.iloc[:,22]
    col24 = df.iloc[:,24]
    mlamtlbl.config(text=col22.sum()+col24.sum())

    #  UBL total 

    col11 = df.iloc[:,11]
    ublcslbl.config(text=col11.sum())

    col12 = df.iloc[:,12]
    ublamtlbl.config(text=col12.sum())

    #  Grand total 

    col13 = df.iloc[:,13]
    gtcslbl.config(text=col13.sum())

    col14 = df.iloc[:,14]
    gtamtlbl.config(text=col14.sum())

    #  Littering total 

    col17 = df.iloc[:,17]
    ltcslbl.config(text=col17.sum())

    col18 = df.iloc[:,18]
    ltamtlbl.config(text=col18.sum())

    #  Smokint total 

    col19 = df.iloc[:,19]
    smcslbl.config(text=col19.sum())

    col20 = df.iloc[:,20]
    smamtlbl.config(text=col20.sum())

######################################################################################################

# Update time function

def update_time():
    global datelbl
    # Get the current date and time
    current_time = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
    
    # Update the label with the current date and time
    datelbl.config(text=current_time)
    
    # Schedule the update_time function to be called again after 1000 milliseconds (1 second)
    datelbl.after(1000, update_time)

################################# Date and copy right pack ########################################

cpframe = tk.LabelFrame(win,)
copyrightlbl = tk.Label(cpframe,text='© 2024 CSDN Technology',font=('New Times Roman',10,'bold'),
                                fg='red',justify=LEFT,padx=200).grid(row=0,column=0)

datelbl = tk.Label(cpframe,font=('New Times Roman',10,'bold'),justify=RIGHT,padx=600,fg='blue')
datelbl.grid(row=0,column=1)    
datelbl.config(text=update_time())
cpframe.pack(pady=2,fill=X)

update_summery()
win.mainloop()  # Main Loop close

######################################################################################################
