import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook

class ExcelApp:
    def __init__(self, master):
        self.master = master
        
        self.file_path = 'STN_PCDO.xlsx'
        self.sheets = ['maindata']
        self.sheet_var = tk.StringVar()
        
        self.select_button = tk.Button(master, text="Select Excel File", command=self.select_file)
        self.select_button.pack(pady=10)
        
        self.radio_frame = tk.Frame(master)
        self.radio_frame.pack(pady=10)
        
        self.update_button = tk.Button(master, text="Update Excel", command=self.update_excel)
        self.update_button.pack(pady=10)
    
    def select_file(self):
        self.file_path = filedialog.askopenfilename(
            title="Open Excel File",
            filetypes=(("Excel Files", "*.xlsx"), ("All Files", "*.*"))
        )
        
        if self.file_path:
            self.load_sheets()
    
    def load_sheets(self,sheet):
        df = pd.read_excel('STN_PCDO.xlsx', sheet_name='maindata')
        workbook = load_workbook(self.file_path)
        sheet = workbook['select_file']
        
        self.sheets = pd.ExcelFile(self.file_path).sheet_names
        for widget in self.radio_frame.winfo_children():
            widget.destroy()
        
        for sheet in self.sheets:
            tk.Radiobutton(self.radio_frame, text=sheet, variable=self.sheet_var, value=sheet).pack(anchor=tk.W)
        
        if self.sheets:
            self.sheet_var.set(self.sheets[0])  # Set default selection
    
    def update_excel(self):
        selected_sheet = self.sheet_var.get()
        if not selected_sheet:
            messagebox.showerror("Error", "No sheet selected!")
            return
        
        try:
            df = pd.read_excel(self.file_path, sheet_name=selected_sheet)
            workbook = load_workbook(self.file_path)
            sheet = workbook[selected_sheet]
            
            # Example tuple and data to be written
            search_tuple = ("example1", "example2")
            data_to_write = "New Data"
            
            for row_idx, row in df.iterrows():
                if tuple(row[:len(search_tuple)]) == search_tuple:
                    col_idx = len(row) + 1  # Assuming you want to write in the next column
                    sheet.cell(row=row_idx+2, column=col_idx, value=data_to_write)
                    break
            
            workbook.save(self.file_path)
            messagebox.showinfo("Success", "Excel file updated successfully!")
        except Exception as e:
            messagebox.showerror("Error", str(e))

# Main part of the program
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelApp(root)
    root.mainloop()
